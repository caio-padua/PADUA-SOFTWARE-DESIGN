# -*- coding: utf-8 -*-
"""Confere se o estado de cada item da fila bate com o estado das dependencias.

Existe porque um estado escrito a mao envelhece calado: o item 31 ficou
marcado como AGUARDANDO depois de o item 7 (a dependencia dele) ficar
CONCLUIDO, e ninguem viu. Um agente lendo aquilo conclui que nao tem
trabalho livre quando tem.

Regra de estado adotada na fila:
  CONCLUIDO   ja entregue e verificado
  A FAZER     nada pendente, ou o pendente e do mesmo dono (fila propria)
  AGUARDANDO  ha pendencia de OUTRO dono na cadeia (nao depende de mim comecar)

A cadeia se le por inteiro, nao so um passo: o item 21 depende do 20, que
depende do 19, que e do Dr. Caio. O 21 espera o Dr. Caio, mesmo sem citar
o Dr. Caio na propria linha. Conferir um passo so daria falso positivo.

Duas divergencias importam, e nenhuma das duas e questao de gosto:
  1. item que espera OUTRO dono mas nao esta marcado AGUARDANDO
     -> o dono comeca e colide, que e o acidente que a fila existe para evitar
  2. item marcado AGUARDANDO com a cadeia inteira limpa
     -> esconde trabalho livre (foi o caso do item 31)
Item travado apenas pela propria fila do mesmo dono nao e erro em nenhum
dos dois estados: e ordem de trabalho dele, e ele decide.

Uso:
  python3 conferir_coerencia.py                                   # nomes padrao
  python3 conferir_coerencia.py --file planilha.xlsx              # arquivo custom
  python3 conferir_coerencia.py --aba "OUTRA ABA"                 # aba custom
  python3 conferir_coerencia.py --file p.xlsx --aba "MINHA ABA"  # ambos

Saida: lista de divergencias e codigo de saida 1, ou "coerente" e 0.
"""

import argparse
import os
import sys

import openpyxl

# --------------------------------------------------------------------------- #
# Argumentos de linha de comando
# --------------------------------------------------------------------------- #
DEFAULT_ARQUIVO = "NOME_DO_PROJETO_LINHA_DE_PRODUCAO_aa.mm.dd.xlsx"
DEFAULT_ABA     = "PLANILHA PRODUCAO"

parser = argparse.ArgumentParser(
    description="Confere coerencia de estados na fila de trabalho (xlsx).")
parser.add_argument(
    "--file",
    default=DEFAULT_ARQUIVO,
    metavar="ARQUIVO.xlsx",
    help="Caminho para o arquivo .xlsx (padrao: %(default)s)")
parser.add_argument(
    "--aba",
    default=DEFAULT_ABA,
    metavar="NOME_DA_ABA",
    help="Nome da aba dentro do arquivo (padrao: '%(default)s')")
args = parser.parse_args()

ORIGEM = args.file
ABA    = args.aba

# --------------------------------------------------------------------------- #
# Validacoes antecipadas — erros claros antes de qualquer KeyError
# --------------------------------------------------------------------------- #
if not os.path.isfile(ORIGEM):
    print("ERRO: arquivo nao encontrado — '{}'".format(ORIGEM))
    print()
    print("  Dica: passe o caminho correto com --file:")
    print("    python3 conferir_coerencia.py --file minha_planilha.xlsx")
    print()
    print("  O nome padrao esperado e:")
    print("    {}".format(DEFAULT_ARQUIVO))
    sys.exit(2)

wb = openpyxl.load_workbook(ORIGEM, data_only=True)

if ABA not in wb.sheetnames:
    print("ERRO: aba '{}' nao encontrada em '{}'".format(ABA, ORIGEM))
    print()
    print("  Abas disponiveis neste arquivo:")
    for nome in wb.sheetnames:
        print("    - {}".format(nome))
    print()
    print("  Dica: passe o nome correto com --aba:")
    print("    python3 conferir_coerencia.py --aba \"NOME DA ABA\"")
    print()
    print("  O nome padrao esperado e:")
    print("    {}".format(DEFAULT_ABA))
    sys.exit(2)

ws = wb[ABA]

# --------------------------------------------------------------------------- #
# Colunas da aba (1-based): Nº | DATA | HORA | ITEM | Responsavel | Estado | Prioridade | Depende de
# --------------------------------------------------------------------------- #
COL_NUMERO, COL_ITEM, COL_DONO, COL_ESTADO, COL_DEPENDE = 1, 4, 5, 6, 8
PRIMEIRA_LINHA = 2   # linha 1 e o cabecalho

# Ultima linha descoberta, nao escrita a mao: com limite fixo a fila cresce
# e o conferidor passa a ignorar o item novo em silencio — que e a mesma
# classe de erro que ele existe para pegar.
# A planilha armazena numeros como float (1.0, 2.0 …); aceitar int e float.
def _is_numero(val):
    return isinstance(val, (int, float)) and not isinstance(val, bool)

ULTIMA_LINHA = PRIMEIRA_LINHA - 1
while _is_numero(ws.cell(ULTIMA_LINHA + 1, COL_NUMERO).value):
    ULTIMA_LINHA += 1

estado, dono, dependencias, titulo = {}, {}, {}, {}
for linha in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
    numero = int(ws.cell(linha, COL_NUMERO).value)
    estado[numero] = ws.cell(linha, COL_ESTADO).value
    dono[numero] = ws.cell(linha, COL_DONO).value
    titulo[numero] = ws.cell(linha, COL_ITEM).value
    bruto = ws.cell(linha, COL_DEPENDE).value
    vazio = bruto in (None, "", "—")
    dependencias[numero] = [] if vazio else [int(float(p)) for p in str(bruto).split(",")]

def ancestrais_pendentes(numero, visitados=None):
    """Toda dependencia ainda nao concluida na cadeia, em qualquer profundidade."""
    visitados = visitados if visitados is not None else set()
    pendentes = set()
    for d in dependencias.get(numero, []):
        if d in visitados:
            continue
        visitados.add(d)
        if estado[d] != "CONCLUIDO":
            pendentes.add(d)
            pendentes |= ancestrais_pendentes(d, visitados)
    return pendentes


divergencias = []

# Numero repetido e o defeito mais perigoso da fila: em 30/07/2026 uma
# segunda fila escrita em paralelo trouxe os numeros 11 e 15 duas vezes,
# cada ocorrencia com item, dono e estado diferentes. Um recado dizendo
# "comece o item 11" virou impossivel de obedecer. Numero de item e
# identidade permanente, igual a versao de documento no Art. 28-A: nunca
# se reaproveita, nunca se reatribui.
vistos = {}
for linha in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
    numero = int(ws.cell(linha, COL_NUMERO).value)
    if numero in vistos:
        divergencias.append(
            "item {} aparece duas vezes (linhas {} e {}): '{}' e '{}'".format(
                numero, vistos[numero], linha,
                titulo[numero], ws.cell(linha, COL_ITEM).value))
    vistos[numero] = linha

for numero in sorted(estado):
    if estado[numero] == "CONCLUIDO":
        continue
    pendentes = ancestrais_pendentes(numero)
    de_outro_dono = sorted(d for d in pendentes if dono[d] != dono[numero])

    if de_outro_dono and estado[numero] != "AGUARDANDO":
        divergencias.append(
            "item {} ({}) marcado {} mas espera o item {} de {}".format(
                numero, dono[numero], estado[numero],
                de_outro_dono[0], dono[de_outro_dono[0]]))
    if not pendentes and estado[numero] == "AGUARDANDO" and dependencias[numero]:
        divergencias.append(
            "item {} ({}) marcado AGUARDANDO com a cadeia limpa — deveria ser A FAZER".format(
                numero, dono[numero]))

if divergencias:
    print("DIVERGENCIAS ({}):".format(len(divergencias)))
    for d in divergencias:
        print("  " + d)
    sys.exit(1)

livres = [n for n in sorted(estado)
          if estado[n] == "A FAZER"
          and all(estado[d] == "CONCLUIDO" for d in dependencias[n])]
print("coerente: {} itens, nenhuma divergencia de estado".format(len(estado)))
print("livres para comecar agora ({}):".format(len(livres)))
for n in livres:
    print("  item {:>2} | {:<16} | {}".format(n, dono[n], titulo[n]))
