# -*- coding: utf-8 -*-
"""Espelho em Markdown da planilha da linha de producao.

Le o .xlsx gerado por gerar.py e escreve o .md que vai para o repositorio.
Motivo de existir: a planilha e para o CEO (Drive); o .md e para os agentes
(Dr. Replit e Dr. Lovable leem texto no repositorio, sem depender de copiar
e colar). Uma fonte, dois formatos — nunca divergem porque este script le
a planilha, nao uma segunda copia dos dados.
"""

import openpyxl

ORIGEM = "NOME_DO_PROJETO_LINHA_DE_PRODUCAO_aa.mm.dd.xlsx"
DESTINO = "FILA_DE_PRODUCAO.md"

wb = openpyxl.load_workbook(ORIGEM, data_only=True)


def celula(ws, r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).replace("|", "\\|")


linhas = []
A = linhas.append

A("# FILA DE PRODUCAO — <NOME DO PROJETO>")
A("")
A("> Fila unica de trabalho. Gerada por script a partir da planilha.")
A("> Existe porque varios agentes trabalham no mesmo sistema sem se ver, e o")
A("> git aceita dois lados do mesmo assunto sem apontar conflito.")
A(">")
A("> Este arquivo e o ESPELHO da planilha de coordenacao do projeto,")
A("> que vive na nuvem. A planilha e para o dono do projeto: ele")
A("> marca estado ali. Este .md e para os agentes lerem sem depender de")
A("> copiar e colar. Quem muda o conteudo muda os dois.")
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 1
ws = wb["LINHA DE PRODUCAO"]

# A ultima linha da fila e descoberta, nao escrita a mao: a fila cresce
# quando entra item novo (o 37 e o 38 entraram em 30/07) e um limite fixo
# passaria a cortar item calado.
PRIMEIRA_LINHA = 5
ULTIMA_LINHA = PRIMEIRA_LINHA - 1
while isinstance(ws.cell(ULTIMA_LINHA + 1, 1).value, int):
    ULTIMA_LINHA += 1
TOTAL = ULTIMA_LINHA - PRIMEIRA_LINHA + 1

A("## 1. A FILA ({} itens, ordenada por prioridade)".format(TOTAL))
A("")
A("Ordem por PRIORIDADE, nao por ordem em que foi pedido. P0 vem antes de")
A("tudo: sao os itens de isolamento entre clinicas — enquanto abertos, uma")
A("segunda clinica no sistema pode ver dado da primeira.")
A("")
cab = [celula(ws, 4, c) for c in range(1, 11)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
for r in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 11)) + " |")
A("")
# Contagem lida das celulas, nunca escrita a mao: numero escrito a mao
# envelhece calado quando um item muda de estado.
estados = [celula(ws, r, 6) for r in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1)]
A("Contagem: {} concluidos, {} a fazer, {} aguardando dependencia.".format(
    estados.count("CONCLUIDO"),
    estados.count("A FAZER"),
    estados.count("AGUARDANDO"),
))
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 2
ws = wb["MATRIZ DA EQUIPE"]
A("## 2. MATRIZ DA EQUIPE — territorio e fronteira")
A("")
cab = [celula(ws, 4, c) for c in range(1, 6)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
r = 5
while celula(ws, r, 1):
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 6)) + " |")
    r += 1
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 3
ws = wb["DECISOES DO DONO"]
A("## 3. DECISOES QUE SO O DONO DO PROJETO TOMA")
A("")
A("Cada linha aqui esta travando trabalho. Nenhum agente decide por ele.")
A("")
cab = [celula(ws, 4, c) for c in range(1, 6)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
r = 5
while celula(ws, r, 2):
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 6)) + " |")
    r += 1
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 4
ws = wb["REGRAS DE EDICAO"]
A("## 4. REGRAS DE EDICAO — valem para todos os agentes")
A("")
A("A regra mestra e uma: NADA SE APAGA, tudo se acrescenta. Sem o antes")
A("continuar existindo, nao se compara o que foi pedido com o que foi feito.")
A("")
cab = [celula(ws, 4, c) for c in range(1, 5)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
r = 5
while celula(ws, r, 1):
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 5)) + " |")
    r += 1
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 5
ws = wb["HISTORICO"]
A("## 5. HISTORICO DE MUDANCAS (append-only)")
A("")
A("So cresce. Linha nunca se apaga nem se corrige: mudou de novo, entra")
A("linha nova. Mais antigo em cima. E aqui que se le como o item estava")
A("antes de cada correcao.")
A("")
cab = [celula(ws, 4, c) for c in range(1, 9)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
r = 5
while celula(ws, r, 1) and celula(ws, r, 1) != "acrescente aqui":
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 9)) + " |")
    r += 1
A("")
A("---")
A("")

# ---------------------------------------------------------------- linha 6
ws = wb["GLOSSARIO"]
A("## 6. GLOSSARIO — toda palavra tecnica que a equipe usa")
A("")
A("Ordem alfabetica automatica, feita pelo script. Palavra nova aparece numa")
A("conversa, vira linha nova aqui. Existe porque o CEO e medico e esta")
A("aprendendo IA, programacao e arquitetura de sistemas.")
A("")
cab = [celula(ws, 4, c) for c in range(1, 6)]
A("| " + " | ".join(cab) + " |")
A("|" + "|".join([" --- "] * len(cab)) + "|")
r = 5
while celula(ws, r, 1) and celula(ws, r, 1) != "acrescente aqui":
    A("| " + " | ".join(celula(ws, r, c) for c in range(1, 6)) + " |")
    r += 1
A("")
A("---")
A("")

# ---------------------------------------------------------------- rodape
A("## 7. COMO USAR")
A("")
A("| Quem | O que faz com este arquivo |")
A("| --- | --- |")
A("| Dono do projeto | Marca `Estado` na planilha e responde a coluna `Sua resposta` das decisoes |")
A("| Escritor unico da fila | Mantem este arquivo e a planilha em sincronia |")
A("| Cada agente | Le a fila, pega item marcado como seu, NAO mexe em item de outro sem avisar |")
A("")
A("Trabalho novo entra como linha nova, com numero maior que o ultimo em uso")
A("(hoje o {} — o proximo e o {}). Numero nunca se reaproveita, nunca se".format(TOTAL, TOTAL + 1))
A("reatribui e nunca se renumera: e identidade permanente, igual a versao de")
A("documento. Uma segunda fila que reatribui numero em uso produz recado")
A("impossivel de obedecer: o mesmo numero passa a significar duas coisas.")
A("")
A("## RESUMO EM 1 LINHA")
A("")
A("Fila unica de {} itens numerados, com dono definido por item, matriz de".format(TOTAL))
A("territorio para os quatro agentes pararem de se atropelar, e as 10")
A("decisoes que estao travando trabalho e so o CEO pode desbloquear.")
A("")

with open(DESTINO, "w", encoding="utf-8") as fh:
    fh.write("\n".join(linhas))

print(f"gerado: {DESTINO} ({len(linhas)} linhas)")
