"""Gera a planilha de coordenacao de um projeto — MODELO.

Copie este arquivo para o repositorio do SEU projeto e troque:

  NOME_DO_PROJETO  pelo nome do projeto
  ITENS            pelos itens reais, descobertos LENDO O CODIGO
  LEIGO            pela traducao de cada item para linguagem de leigo
  EQUIPE           pelos agentes reais e o territorio de cada um
  DECISOES         pelas decisoes que so o dono pode responder
  GLOSSARIO        pelas palavras tecnicas que aparecerem no projeto

Os tres itens de exemplo existem so para mostrar os tres estados possiveis.
Apague-os.

AVISO DE SEGURANCA: nunca escreva, em repositorio publico, qual falha esta
ABERTA hoje num sistema em producao. Descreva a categoria ("isolamento entre
clientes"), nunca o estado atual. Repositorio publico que enumera falha
aberta e mapa para quem quer atacar aquele sistema.
"""
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NOME_DO_PROJETO = "NOME DO PROJETO"
NOME_DO_DONO = "DONO"
SAIDA = "NOME_DO_PROJETO_LINHA_DE_PRODUCAO_aa.mm.dd.xlsx"

# ---------------------------------------------------------------- estilo
FONTE = "Arial"
TINTA = "2A231B"
OURO = "9A6B28"
CABEC_FILL = PatternFill("solid", fgColor="2A231B")
TITULO_FILL = PatternFill("solid", fgColor="F2E6D0")
PASTEL = PatternFill("solid", fgColor="F6F2EA")
FEITO = PatternFill("solid", fgColor="E2EBDE")
ESPERA = PatternFill("solid", fgColor="F4EAD2")
AMARELO = PatternFill("solid", fgColor="FFFFCC")

fina = Side(style="thin", color="DDD2BF")
BORDA = Border(left=fina, right=fina, top=fina, bottom=fina)


def cabecalho(ws, linha, textos, larguras):
    for i, (t, w) in enumerate(zip(textos, larguras), start=1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = Font(name=FONTE, size=9, bold=True, color="FFFFFF")
        c.fill = CABEC_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 30


def titulo(ws, texto, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = Font(name=FONTE, size=15, bold=True, color=TINTA)
    c.fill = TITULO_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=sub)
    c.font = Font(name=FONTE, size=9, italic=True, color="6B5F4E")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def corpo(ws, linha, valores, fill=None):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = Font(name=FONTE, size=9, color=TINTA)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = BORDA
        if fill:
            c.fill = fill


def chave_alfabetica(termo):
    """Ordem alfabetica do portugues: acento nao joga a palavra para o fim.

    Sem isso 'indice' cai depois de 'webhook', porque o computador ordena
    pelo numero interno da letra e a letra acentuada tem numero alto.
    """
    sem_acento = unicodedata.normalize("NFKD", termo)
    return "".join(c for c in sem_acento if not unicodedata.combining(c)).lower()


wb = Workbook()

# ================================================================ ABA 1
ws = wb.active
ws.title = "LINHA DE PRODUCAO"
COLS = ["Nº", "Item", "O que e / como sera feito", "Responsavel",
        "Depende de", "Estado", "Prioridade", "Esforco", "Observacao",
        "Em linguagem de leigo"]
LARG = [5, 30, 52, 14, 12, 14, 11, 10, 40, 58]
titulo(ws, f"{NOME_DO_PROJETO} — Linha de Producao",
       "Fila unica de trabalho. Ordenada por prioridade, nao por ordem de pedido. "
       "Edite apenas as colunas Estado e Observacao.",
       len(COLS))
cabecalho(ws, 4, COLS, LARG)

# (numero, item, descricao, responsavel, depende, estado, prioridade, esforco, observacao)
#
# APAGUE estes tres e ponha os itens reais, descobertos LENDO O CODIGO.
# Fila montada a partir de pergunta ao dono nasce com os itens que ele
# lembrou; montada a partir do codigo, nasce com os itens que existem.
ITENS = [
    (1, "Exemplo de item concluido",
     "Descricao concreta: nome do arquivo, da tabela ou da rota.",
     "Agente A", "—", "CONCLUIDO", "P0 critico", "feito",
     "Prova obrigatoria: identificador do commit e arquivo tocado."),
    (2, "Exemplo de item livre para comecar",
     "Nada pendente na cadeia, ou o pendente e do mesmo dono.",
     "Agente A", "—", "A FAZER", "P1 alto", "3h",
     "Qualquer agente escreve nesta coluna, assinando com nome e data."),
    (3, "Exemplo de item travado por outro agente",
     "Depende do item 2, que e de outro dono.",
     "Agente B", "2", "AGUARDANDO", "P2 medio", "1 dia",
     "AGUARDANDO significa que a pendencia e de OUTRO dono na cadeia."),
]

# Traducao de cada item para linguagem de leigo, pela chave do numero.
# Obrigatoria quando o dono do projeto nao e programador: cada sigla nao
# explicada e uma decisao que ele nao consegue tomar.
#
# A analogia sai do mundo DELE (clinica, processo, obra), nunca do mundo da
# computacao. E tem de ser honesta: analogia bonita que ensina errado e pior
# do que nenhuma, porque ele vai decidir com base nela.
LEIGO = {
    1: "Exemplo de traducao: a analogia vem do mundo do dono do projeto.",
    2: "Exemplo: 'e a tranca da porta, nao o aviso pedindo para nao entrar'.",
    3: "Exemplo: 'espera o andar de baixo ficar pronto para subir a parede'.",
}

for k, it in enumerate(ITENS):
    linha = 5 + k
    estado = it[5]
    fill = FEITO if estado == "CONCLUIDO" else (ESPERA if estado == "AGUARDANDO" else None)
    corpo(ws, linha, list(it) + [LEIGO.get(it[0], "")], fill)
    ws.row_dimensions[linha].height = 46

ULT = 4 + len(ITENS)

dv_estado = DataValidation(type="list",
                           formula1='"A FAZER,AGUARDANDO,CONCLUIDO"', allow_blank=True)
dv_prio = DataValidation(type="list",
                         formula1='"P0 critico,P1 alto,P2 medio,P3 baixo"', allow_blank=True)
ws.add_data_validation(dv_estado)
ws.add_data_validation(dv_prio)
dv_estado.add(f"F5:F{ULT}")
dv_prio.add(f"G5:G{ULT}")
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:J{ULT}"

# Resumo por formula: numero calculado nunca envelhece calado. Numero escrito
# a mao continua dizendo 36 quando ja sao 38.
RESUMO = [("Concluido", "CONCLUIDO"), ("A fazer", "A FAZER"),
          ("Aguardando dependencia", "AGUARDANDO"), ("Total de itens", None)]
ws.cell(row=ULT + 2, column=1, value="RESUMO").font = Font(
    name=FONTE, size=10, bold=True, color=OURO)
for i, (rotulo, alvo) in enumerate(RESUMO):
    lr = ULT + 3 + i
    ws.cell(row=lr, column=1, value=rotulo).font = Font(name=FONTE, size=9, color=TINTA)
    f = (f'=COUNTIF(F5:F{ULT},"{alvo}")' if alvo else f"=COUNTA(A5:A{ULT})")
    ws.cell(row=lr, column=2, value=f).font = Font(name=FONTE, size=9, bold=True, color=TINTA)

# ================================================================ ABA 2
ws2 = wb.create_sheet("MATRIZ DA EQUIPE")
COLS2 = ["Membro", "Onde vive", "Territorio (decide sozinho)",
         "NAO toca sem avisar", "Frente atual"]
LARG2 = [24, 18, 46, 40, 34]
titulo(ws2, f"{NOME_DO_PROJETO} — Matriz da equipe",
       "A coluna que evita a colisao e a quarta: o que cada um NAO toca sem avisar. "
       "A linha do escritor unico da fila vem marcada.",
       len(COLS2))
cabecalho(ws2, 4, COLS2, LARG2)

EQUIPE = [
    (NOME_DO_DONO, "—", "Preco, produto, prioridade, visao, risco aceitavel",
     "Nada: decide tudo", "As decisoes em aberto"),
    ("Agente A (escritor da fila)", "plataforma",
     "Servidor, banco, infraestrutura, coordenacao da fila",
     "Telas e site", "itens 1 e 2"),
    ("Agente B", "plataforma", "Telas do sistema, logica de negocio",
     "Banco em producao, configuracao de infraestrutura", "item 3"),
]
for i, e in enumerate(EQUIPE):
    corpo(ws2, 5 + i, e, PASTEL if i % 2 == 0 else None)
    ws2.row_dimensions[5 + i].height = 40
ws2.freeze_panes = "A5"

# ================================================================ ABA 3
ws3 = wb.create_sheet("DECISOES DO DONO")
COLS3 = ["Nº", "Decisao", "Por que so voce decide", "Trava o item", "Sua resposta"]
LARG3 = [5, 44, 46, 14, 44]
titulo(ws3, f"{NOME_DO_PROJETO} — Decisoes que so o dono toma",
       "Cada linha aqui esta travando trabalho. Nenhum agente decide por ele. "
       "Escreva na ultima coluna, em amarelo.",
       len(COLS3))
cabecalho(ws3, 4, COLS3, LARG3)

DECISOES = [
    (1, "Exemplo: valores em reais de cada modulo",
     "Preco nasce do dono. Nenhum agente decide por ele.", "2,3", ""),
]
for i, d in enumerate(DECISOES):
    corpo(ws3, 5 + i, d, PASTEL if i % 2 == 0 else None)
    ws3.cell(row=5 + i, column=5).fill = AMARELO
    ws3.row_dimensions[5 + i].height = 40
ws3.freeze_panes = "A5"

# ================================================================ ABA 4
ws4 = wb.create_sheet("REGRAS DE EDICAO")
COLS4 = ["Regra", "O que pode", "O que nao pode", "Por que"]
LARG4 = [26, 44, 44, 52]
titulo(ws4, f"{NOME_DO_PROJETO} — Regras de edicao desta planilha",
       "Vale para o dono e para todos os agentes. "
       "A regra mestra e uma: NADA SE APAGA, tudo se acrescenta.",
       len(COLS4))
cabecalho(ws4, 4, COLS4, LARG4)

REGRAS = [
    ("Nada se apaga",
     "Acrescentar linha, observacao, comentario.",
     "Apagar linha, apagar texto de outra pessoa, sobrescrever celula preenchida.",
     "So se compara o que foi pedido com o que foi feito se o antes continuar existindo."),
    ("Numero de item e permanente",
     "Criar item novo com numero maior que o ultimo em uso.",
     "Reaproveitar, reatribuir ou renumerar numero existente.",
     "Todo recado antigo que cita numero passaria a apontar para o item errado."),
    ("Mudanca de estado vira linha no HISTORICO",
     "Mudar Estado e registrar: data, quem, de que para que, e por que.",
     "Mudar Estado sem deixar rastro.",
     "Sem rastro ninguem sabe se o item foi concluido, corrigido ou mexido por engano."),
    ("Concluido exige prova",
     "Marcar CONCLUIDO citando o identificador do commit e o arquivo tocado.",
     "Marcar CONCLUIDO por prosa, memoria ou intencao.",
     "Anunciar sem prova destroi a confianca na fila inteira."),
    ("Cada um na sua coluna",
     "Qualquer agente escreve em Observacao, assinando com nome e data.",
     "Um agente mudar Responsavel, Depende de ou Prioridade de item de outro.",
     "Territorio esta na aba MATRIZ DA EQUIPE."),
    ("Um escritor da estrutura",
     "Pedir mudanca de estrutura pela Observacao ou pelo recado.",
     "Dois agentes gerando a planilha ou subindo arquivo de mesmo nome.",
     "Recurso compartilhado com dois escritores corrompe em silencio."),
    ("As duas copias dizem a mesma coisa",
     "Ler na nuvem (o dono) ou no repositorio (os agentes).",
     "Editar so uma das duas e deixar a outra velha.",
     "O espelho em texto e gerado desta planilha por script. Divergiu = alguem digitou."),
    ("Decisao do dono e so do dono",
     "Qualquer um pode registrar uma pergunta na aba de decisoes.",
     "Qualquer agente responder no lugar dele sobre preco, produto ou risco.",
     "Nao e territorio de agente nenhum."),
]
for i, r in enumerate(REGRAS):
    corpo(ws4, 5 + i, r, PASTEL if i % 2 == 0 else None)
    ws4.row_dimensions[5 + i].height = 58
ws4.freeze_panes = "A5"

# ================================================================ ABA 5
ws5 = wb.create_sheet("HISTORICO")
COLS5 = ["Data", "Quem", "Nº", "Campo", "Como estava", "Como ficou",
         "Por que mudou", "Prova"]
LARG5 = [11, 13, 5, 14, 20, 20, 50, 26]
titulo(ws5, f"{NOME_DO_PROJETO} — Historico de mudancas (append-only)",
       "So cresce. Linha nunca se apaga nem se corrige: mudou de novo, entra linha nova. "
       "Mais antigo em cima.",
       len(COLS5))
cabecalho(ws5, 4, COLS5, LARG5)

HISTORICO = [
    ("dd/mm/aaaa", "Agente A", "—", "criacao", "—", "3 itens",
     "Primeira versao da fila, montada a partir da leitura do codigo.",
     "commit <identificador>"),
]
for i, h in enumerate(HISTORICO):
    corpo(ws5, 5 + i, h, PASTEL if i % 2 == 0 else None)
    ws5.row_dimensions[5 + i].height = 46
ws5.freeze_panes = "A5"
ws5.auto_filter.ref = f"A4:H{4 + len(HISTORICO)}"
PROX5 = 5 + len(HISTORICO)
c = ws5.cell(row=PROX5, column=1, value="acrescente aqui")
c.font = Font(name=FONTE, size=9, italic=True, color="9A8F7E")
for col in range(1, len(COLS5) + 1):
    ws5.cell(row=PROX5, column=col).fill = AMARELO
    ws5.cell(row=PROX5, column=col).border = BORDA

# ================================================================ ABA 6
ws6 = wb.create_sheet("GLOSSARIO")
COLS6 = ["Palavra", "Sinonimos e como aparece", "O que e",
         "Em linguagem de leigo, com analogia", "Para que serve"]
LARG6 = [20, 26, 46, 62, 46]
titulo(ws6, f"{NOME_DO_PROJETO} — Glossario para o dono do projeto",
       "Toda palavra tecnica usada pela equipe entra aqui. Ordem alfabetica automatica. "
       "Palavra nova aparece numa conversa, vira linha nova aqui.",
       len(COLS6))
cabecalho(ws6, 4, COLS6, LARG6)

# Quatro exemplos, para mostrar o padrao da analogia honesta. Acrescente as
# palavras que aparecerem no SEU projeto. A coluna "para que serve" amarra o
# termo a um item real da fila: termo sem uso concreto nao gruda.
GLOSSARIO = [
    ("append-only", "so cresce, log append-only",
     "Registro em que se acrescenta linha, nunca se apaga nem se corrige linha antiga.",
     "E o prontuario do paciente: nao se apaga a evolucao de ontem, escreve-se a de hoje embaixo. Se errou ontem, registra-se hoje que errou. O erro fica, e e isso que da valor legal.",
     "A aba HISTORICO e append-only. E o que permite comparar pedido com entrega."),
    ("commit", "gravacao, ponto de salvamento",
     "Um pacote de mudancas no codigo, com autor, data e explicacao, gravado de forma permanente.",
     "E uma entrada assinada no prontuario: quem escreveu, quando, e o que mudou. Nao se apaga.",
     "E a prova exigida para marcar um item como concluido."),
    ("hash", "resumo criptografico, impressao digital, identificador de commit",
     "Numero unico calculado a partir de um conteudo. Mudou uma virgula, o numero muda inteiro.",
     "E a impressao digital de um documento: unica, e nao bate mais se o documento foi adulterado.",
     "Serve para provar entrega. Sem ele, 'esta pronto' e afirmacao, nao fato."),
    ("idempotente", "reprocessavel, seguro para repetir",
     "Operacao que pode ser repetida quantas vezes for e o resultado final e sempre o mesmo.",
     "E apertar o botao do elevador: apertar dez vezes nao chama dez elevadores.",
     "Toda operacao que grava dado tem de poder rodar de novo sem duplicar."),
]
for i, g in enumerate(sorted(GLOSSARIO, key=lambda x: chave_alfabetica(x[0]))):
    corpo(ws6, 5 + i, g, PASTEL if i % 2 == 0 else None)
    ws6.row_dimensions[5 + i].height = 56
ws6.freeze_panes = "A5"
ws6.auto_filter.ref = f"A4:E{4 + len(GLOSSARIO)}"
PROX6 = 5 + len(GLOSSARIO)
c = ws6.cell(row=PROX6, column=1, value="acrescente aqui")
c.font = Font(name=FONTE, size=9, italic=True, color="9A8F7E")
for col in range(1, len(COLS6) + 1):
    ws6.cell(row=PROX6, column=col).fill = AMARELO
    ws6.cell(row=PROX6, column=col).border = BORDA

wb.save(SAIDA)
print("gerado:", SAIDA)
print("itens:", len(ITENS), "| equipe:", len(EQUIPE),
      "| decisoes:", len(DECISOES), "| glossario:", len(GLOSSARIO))
print()
print("PROXIMO PASSO: troque os itens de exemplo pelos itens reais do projeto,")
print("descobertos LENDO O CODIGO, e rode conferir_coerencia.py.")
